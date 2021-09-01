import smtplib
import openpyxl
import sys

wb = openpyxl.load_workbook("duesRecords.xlsx")
sheet = wb["Sheet1"]
last_col = sheet.max_column
last_month = sheet.cell(row=1,column=last_col).value

unpaid_members = {}

for r in range(2,sheet.max_row+1):
    payment = sheet.cell(row=r,column=last_col).value
    if payment != "paid":
        name = sheet.cell(row=r,column=1).value
        email = sheet.cell(row=r,column=2).value
        unpaid_members[name] = email

print(unpaid_members)

server = smtplib.SMTP_SSL("smtp.gmail.com",587)
server.ehlo()
server.starttls()
server.login("jakis@gmail.com","jakieshaslo")

for name,email in unpaid_members.items():
    body = f"Subject: {last_month} dues unpaid\n Dear {name}, reports show that you have not paid dues for {last_month}.Please make payment as soon as possible. Thank You!"
    print(f"Sending email to {name}....")
    send_mail_status = server.sendmail("jakis@gmail.com",email,body)
    if send_mail_status != {}:
        print(f"There was a problem with sending an email to {name} ")
    server.quit()


