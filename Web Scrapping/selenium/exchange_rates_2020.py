import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.support.ui import Select


start = time.time()
exchange_rates = {}
url = "https://www.nbp.pl/home.aspx?c=/ascx/archa.ascx"

def data_scrapper(url,dict):
    browser = webdriver.Firefox()
    browser.get(url)
    accept = browser.find_element_by_css_selector(".cbutton")
    accept.click()
    select = Select(browser.find_element_by_name("rok"))
    select.select_by_visible_text("2020")


    for index in range(12):
        element = Select(browser.find_element_by_name("mies"))
        options = element.options
        el = options[index]
        element.select_by_index(index)
        el.click()
        submit = browser.find_element_by_css_selector("#ctl25_btAscxSubmit")
        submit.click()


        daysEl = browser.find_elements_by_partial_link_text("Tabela nr")


        for index in range(len(daysEl) - 1):
            time.sleep(1)
            daysEl = browser.find_elements_by_partial_link_text("Tabela nr")
            date = (daysEl[index].text.split()[-1])
            el = daysEl[index]
            el.click()
            dolar = browser.find_element_by_css_selector(".pad5 > tbody:nth-child(2) > tr:nth-child(3) > td:nth-child(3)")
            euro = browser.find_element_by_css_selector(".pad5 > tbody:nth-child(2) > tr:nth-child(9) > td:nth-child(3)")
            funt = browser.find_element_by_css_selector(".pad5 > tbody:nth-child(2) > tr:nth-child(12) > td:nth-child(3)")
            dict[date] = {"dolar": dolar.text, "euro": euro.text,"funt": funt.text}
            browser.back()

        browser.back()
    return dict

def save_to_excel(dict,filename):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Exchange rates 2020"
    sheet["A1"] = "Date"
    sheet["B1"] = "Dolar"
    sheet["C1"] = "Euro"
    sheet["D1"] = "Funt"

    index = 0
    for row in range(2, len(exchange_rates) + 1):
        date = list(exchange_rates.keys())[index]
        sheet.cell(row=row,column=1).value = date
        sheet.cell(row=row,column=2).value = dict[date]["dolar"]
        sheet.cell(row=row, column=3).value = dict[date]["euro"]
        sheet.cell(row=row, column=4).value = dict[date]["funt"]
        index += 1

    wb.save(filename)

data_scrapper(url,exchange_rates)
save_to_excel(exchange_rates,"exchange_rates_2020.xlsx")

stop = time.time()
working_time = stop - start
print(working_time)
