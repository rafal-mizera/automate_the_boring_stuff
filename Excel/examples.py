import openpyxl
from openpyxl.cell.cell import get_column_letter

wb = openpyxl.load_workbook("example.xlsx")
print(type(wb))
print(wb.sheetnames)
sheet = wb["Sheet1"]
# print(sheet)
print(sheet.title)
# anotherSheet = wb.active
# print(anotherSheet)
# print(sheet["A1"])
# print(sheet["A1"].value)
# c = sheet["B1"]
# print(c.value)
# print(f"Row, {c.row} Column {c.column} is {c.value}")
# print(sheet.cell(row = 1, column = 2).value)
# for i in range(1,8,2):
#     print(i,sheet.cell(row = i, column = 2).value)
# print(sheet.max_row)
# print(sheet.max_column)
#
# print(get_column_letter(900))
#
# tuple = tuple(sheet["A1":"C3"])
#
# for row in sheet["A1":"C3"]:
#     for cell in row:
#         print(f"{cell.coordinate} : {cell.value}")
#     print("END OF ROW")
#
# for obj in sheet["A"]:
#     print(obj.coordinate)
###############################################################################
# wb = openpyxl.Workbook()
# sheet  = wb.active
#
# sheet.title = "Spam bacon eggs Sheet"
# print(wb.active)

# wb = openpyxl.load_workbook("example.xlsx")
# sheet = wb.active
# sheet.title = "Spam Spam Spam"
# wb.save("example_copy.xlsx")

# wb = openpyxl.Workbook()
# wb.create_sheet()
# print(wb.sheetnames)
# wb.create_sheet(index=0,title="First Sheet")
# print(wb.sheetnames)
# wb.remove(wb["Sheet1"])
# print(wb.sheetnames)
# sheet = wb.active
# sheet["A1"] = "Hello World!"
# print(sheet["A1"].value)

# from openpyxl.styles import Font, NamedStyle
#
#
# wb = openpyxl.Workbook()
# sheet = wb["Sheet"]
# italic24Font = Font(size=24,italic=True)
# styleObj = NamedStyle(font=italic24Font,name="styleObj")
# styleObj2 = NamedStyle(name="styleObj2")
# styleObj2.font = Font(size=25,bold=True)
# sheet["A1"].style = styleObj
# sheet["A2"].style = styleObj2
# sheet["A1"] = "Hello World"
# sheet["A2"] = "Hello Universe"
#
#
#
#
# wb.save("styled.xlsx")
#########################################################################################
# wb = openpyxl.Workbook()

# sheet = wb.active
# sheet["A1"] = 200
# sheet["A2"] = 300
# sheet["A3"] = '=SUM(A1:A2)'
# # print(sheet["A3"].value)
# wb.save("writeFormula.xlsx")
#
# wbFormulas = openpyxl.load_workbook("writeFormula.xlsx")
# sheet1 = wbFormulas.active
# print(sheet1["A3"].value)
#
# wbDataOnly = openpyxl.load_workbook("writeFormula.xlsx",data_only=True)
# sheet2 = wbDataOnly.active
# print(sheet2["A3"].value)

#####################################################################################
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet["A1"] = "Tall row"
# sheet["B2"] = "Wide column"
# sheet.row_dimensions[1].height = 70
# sheet.column_dimensions["B"].width = 20
# sheet.merge_cells("C1:D3")
# sheet["C1"] = "Merged cells"
#
# wb.save("dimensions.xlsx")
#
# wb = openpyxl.load_workbook("dimensions.xlsx")
# sheet = wb.active
# sheet.unmerge_cells("C1:D3")
#
# wb.save("dimensions2.xlsx")

# wb = openpyxl.load_workbook("produceSales.xlsx")
# sheet = wb.active
# sheet.freeze_panes = "A2"
# wb.save("freezePanes.xlsx")

wb = openpyxl.Workbook()
sheet = wb.active

for i in range(1,11):
    sheet["A" + str(i+1)] = i * 2

chart = openpyxl.chart.BarChart()
chart.type = "col"
chart.style = 10
chart.title = "SampleChart"
chart.y_axis.title = "random"
chart.x_axis.title = "numbers"
data = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_row=11, max_col=1)

chart.add_data(data,titles_from_data=True)
chart.shape = 4
sheet.add_chart(chart,"C1")
wb.save("charts.xlsx")

