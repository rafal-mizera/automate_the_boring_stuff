import openpyxl

wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb["Sheet"]

price_updates = {"Garlic": 3.07, "Celery": 1.19, "Lemon": 1.27}

for row in range(2,sheet.max_row + 1):
    product = sheet.cell(row=row, column=1).value
    if product in price_updates:
        sheet.cell(row=row, column=2).value = price_updates[product]


wb.save("UpdatedProduceSales.xlsx")