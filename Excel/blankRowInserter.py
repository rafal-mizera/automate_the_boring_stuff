import openpyxl
from openpyxl.cell.cell import get_column_letter

filename = "censuspopdata.xlsx"
N = int(input("Input a number: "))
M = int(input("Input a number: "))

wb = openpyxl.load_workbook(filename)
sheet = wb.active
wb2 = openpyxl.Workbook()
sheet2 = wb2.active

for row in sheet["A1": get_column_letter(sheet.max_column) + str(N)]:
    for cell in row:
        sheet2[cell.coordinate] = cell.value

sheet2.insert_rows(M)

for row in sheet["A" + str(N+M+1): get_column_letter(sheet.max_column) + str(sheet.max_row)4]:
    for cell in row:
        sheet2[cell.coordinate] = cell.value


wb2.save("cenuspopdata_copy2.xlsx")


