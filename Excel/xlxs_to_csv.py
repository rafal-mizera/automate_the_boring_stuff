import os
import openpyxl
from openpyxl.cell.cell import get_column_letter
import csv

os.makedirs("csv_copies",exist_ok=True)


for file in os.listdir("."):
    if not file.endswith(".xlsx"):
        continue
    wb = openpyxl.load_workbook(file)
    for sheet in wb.sheetnames:
        print(sheet)
        sheet = wb[sheet]
        with open(os.path.join("csv_copies",f"{os.path.splitext(file)[0]}" + "_" + f"{sheet.title}") + ".csv","w",newline="") as csv_file:
            csv_writer = csv.writer(csv_file)
            for row in range(1,sheet.max_row):
                print(f"Writing row: {row}")
                list = []
                for column in range(1,sheet.max_column):
                    cell = sheet[get_column_letter(column) + str(row)].value
                    list.append(cell)
                csv_writer.writerow(list)




