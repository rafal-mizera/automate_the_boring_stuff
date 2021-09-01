import openpyxl
from openpyxl.cell.cell import get_column_letter

filename = "produceSales.xlsx"

wb = openpyxl.load_workbook(filename)
sheet = wb.active


for column in range(1,sheet.max_column):
    with open(f"new_text_file{column}.txt", "a") as file:
        for cell in sheet[get_column_letter(column)]:
            file.write(str(cell.value))
            file.write("\n")




