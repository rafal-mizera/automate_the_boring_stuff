import openpyxl
import pprint

wb = openpyxl.load_workbook("excel_exchange_rates.xlsx")
sheet = wb.active
exchange_data = {}

for row in range(2,sheet.max_row + 1):
    date = sheet.cell(row=row,column=1).value
    usd = sheet.cell(row=row,column=2).value
    gbp = sheet.cell(row=row,column=3).value
    eur = sheet.cell(row=row,column=4).value
    exchange_data.setdefault(date,{"USD": usd, "GBP": gbp, "EUR": eur})

def show_rate():
    try:
        date = input("Podaj datę w formacie RRRR-MM-DD: ")
        currency = input("Wybierz walutę(USD/GBP/EUR): ")

        rate_from_day = exchange_data[date][currency]

        print(f"Kurs {currency} z dnia {date} wynosił {rate_from_day}")

    except KeyError:
        print("Wprowadź poprawne dane!")
        show_rate()

show_rate()


