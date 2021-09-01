#! python3
# readCensusExcel.py - Tabulates population and number of census tracts for
# each county.

import openpyxl
import pprint

print("Opening workboook...")

wb = openpyxl.load_workbook("censuspopdata.xlsx")
sheet = wb.active
county_data = {}

print("Reading rows...")

for row in range(2,sheet.max_row + 1):
    state = sheet.cell(row=row,column = 2).value
    county = sheet.cell(row=row,column = 3).value
    pop = sheet.cell(row=row,column = 4).value

    county_data.setdefault(state,{})
    county_data[state].setdefault(county,{"tracts": 0, "pop": 0})
    county_data[state][county]["tracts"] += 1
    county_data[state][county]["pop"] += int(pop)

print("Writing results...")
with open("census2010.py", "w") as file:
    file.write("allData =" + pprint.pformat(county_data))

print("Done")

##################################################################3
from automate_boring_stuff.Excel import census2010

print(census2010.allData["AK"]["Anchorage"])






