"""Takes the N number from input and generates NxN multiplication table in new workbook in Excel """
import openpyxl
from openpyxl.cell.cell import get_column_letter
from openpyxl.styles import Font, NamedStyle

try:
    n = int(input("Input number: "))
except ValueError:
    print("You need to input a number...")

wb = openpyxl.Workbook()
sheet = wb.active
Bold = Font(bold=True)



for number in range(1,n+1):
    sheet[get_column_letter(number+1)+str(1)] = number
    sheet[get_column_letter(number + 1) + str(1)].font = Bold
    sheet["A" + str(number+1)] = number
    sheet["A" + str(number + 1)].font = Bold

for row in range(2, n + 2):
    for column in range(2, n + 2):
        sheet[get_column_letter(column)+str(row)] = sheet[get_column_letter(column)+ str(1)].value * sheet["A" + str(row)].value


wb.save("multiplicationTable.xlsx")



